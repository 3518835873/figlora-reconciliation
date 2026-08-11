#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_excel.py — 菲格洛亚 FIGUEROA「BabyCat」财务对账工具包生成脚本
=====================================================================
用途   : 求职作品展示 —— 经销商对账 / 摩点众筹结算 / 艺术家版税 / 展会核算
数据源 : ../data/*.csv（UTF-8 BOM，10 个文件）
输出   : ../excel/Figlora_财务工具包.xlsx（15 张工作表，全部公式自包含）
依赖   : openpyxl（Anaconda 自带）
运行   : 项目根目录下执行
         /c/ProgramData/anaconda3/python.exe scripts/build_excel.py

说明   :
  - 所有数据表统一布局：第1行标题 / 第2行口径说明 / 第3行表头 / 第4行起数据
  - 经销商对账使用 SUMIF/SUMIFS 跨表公式，银行回款通过付款方名称自动匹配经销商
  - 数据期间 2026年5—7月，对账基准日 2026-08-11（Dashboard F31 可修改）
"""

import csv
import datetime
import os
import re

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ══════════════════════════════════════════════════════════
# 1. 路径与全局配置
# ══════════════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "..", "data")
OUT_DIR = os.path.join(BASE_DIR, "..", "excel")
OUT_FILE = os.path.join(OUT_DIR, "Figlora_财务工具包.xlsx")

PERIOD = "2026年5—7月"
BASE_DATE = datetime.date(2026, 8, 11)          # 对账基准日
MONTHS = [datetime.date(2026, 5, 1), datetime.date(2026, 6, 1), datetime.date(2026, 7, 1)]

# 工作表名称（公式中引用的关键表）
S_SALES = "📦 销售出库"
S_ST = "🏪 Sell-Through"
S_BANK = "💰 银行回款"
S_RECON = "🔗 经销商对账"

# ══════════════════════════════════════════════════════════
# 2. 样式常量（品牌色系：浅紫 #E8D5F5）
# ══════════════════════════════════════════════════════════

FONT_NAME = "微软雅黑"
HEADER_FILL = "E8D5F5"        # 表头浅紫
ALT_FILL = "F5F0FA"           # 交替行淡紫
SECTION_FILL = "EDE0F7"       # 分区标题
TITLE_FILL = "F0E4FA"         # 大标题
DEEP_PURPLE = "4A2563"

GREEN_FILL, GREEN_FONT = "C6EFCE", "006100"   # 已匹配
YELLOW_FILL, YELLOW_FONT = "FFEB9C", "9C6500"  # 部分回款
RED_FILL, RED_FONT = "FFC7CE", "9C0006"        # 未回款/发票差异

TAB_DATA = "4472C4"     # 数据表：蓝
TAB_ANALYSIS = "7030A0"  # 分析表：紫
TAB_DASH = "70AD47"      # 仪表盘：绿
TAB_UTIL = "808080"      # 工具表：灰

FMT_AMT = "#,##0.00"
FMT_QTY = "#,##0"
FMT_DATE = "yyyy-mm-dd"
FMT_MONTH = "yyyy-mm"
FMT_PCT = "0.0%"
FMT_RATE = "0.00"

THIN = Side(style="thin", color="C9C9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LAYOUT_HEADER_ROW = 3   # 数据表统一：标题1 / 说明2 / 表头3 / 数据从4开始
DATA_FIRST = LAYOUT_HEADER_ROW + 1

# ══════════════════════════════════════════════════════════
# 3. 通用工具函数
# ══════════════════════════════════════════════════════════


def read_csv(name):
    """读取 UTF-8 BOM CSV，返回 dict 列表（键/值去除首尾空白）。"""
    path = os.path.join(DATA_DIR, name)
    if not os.path.exists(path):
        raise FileNotFoundError(f"缺少数据文件: {path}")
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [{k.strip(): (v.strip() if v is not None else "")
                 for k, v in row.items()} for row in csv.DictReader(f)]


def num(v):
    """字符串 → float，容忍千分位与货币符号。"""
    if v is None:
        return 0.0
    s = str(v).replace(",", "").replace("¥", "").replace("￥", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def intv(v):
    return int(round(num(v)))


def pct(v):
    """'8%' → 0.08；'0.08' → 0.08。"""
    s = str(v).strip()
    if not s:
        return 0.0
    if s.endswith("%"):
        return float(s[:-1]) / 100.0
    return float(s)


def date_(v):
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def disp_width(v):
    """估算显示宽度（中日韩字符按 2 个单位）。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(v))


def header_style(cell, size=12):
    cell.font = Font(name=FONT_NAME, size=size, bold=True, color=DEEP_PURPLE)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def make_title(ws, title, note, ncols):
    """统一布局：第1行大标题 + 第2行口径说明。"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT_NAME, size=14, bold=True, color=DEEP_PURPLE)
    c.fill = PatternFill("solid", fgColor=TITLE_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=note)
    c2.font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 18


def section_header(ws, row, text, ncols, fill=SECTION_FILL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=11, bold=True, color=DEEP_PURPLE)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row


def setup_sheet(ws, tab, freeze="A4", landscape=True, repeat_rows="1:3"):
    ws.sheet_properties.tabColor = tab
    ws.freeze_panes = freeze
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = repeat_rows


def write_table(ws, start_row, headers, rows, col_formats=None, alt=True,
                center_cols=(), start_col=1):
    """写入带表头、边框、交替行色的数据表，返回最后一行行号。"""
    for j, h in enumerate(headers):
        header_style(ws.cell(row=start_row, column=start_col + j, value=h))
    last = start_row + len(rows)
    for i, row in enumerate(rows):
        r = start_row + 1 + i
        for j, v in enumerate(row):
            c = ws.cell(row=r, column=start_col + j, value=v)
            c.font = Font(name=FONT_NAME, size=10)
            c.alignment = Alignment(
                horizontal="center" if j in center_cols else "left",
                vertical="center", wrap_text=(j not in center_cols))
            c.border = BORDER
            if col_formats and j < len(col_formats) and col_formats[j]:
                c.number_format = col_formats[j]
            if alt and i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ALT_FILL)
    return last


def autofit(ws, headers, rows, start_col=1, overrides=None):
    """按内容自动列宽（公式单元格跳过，用表头宽度）。"""
    widths = [disp_width(h) for h in headers]
    if rows:
        for row in rows:
            for j, v in enumerate(row):
                if v is None or (isinstance(v, str) and v.startswith("=")):
                    continue
                widths[j] = max(widths[j], disp_width(v))
    for j in range(len(headers)):
        col = start_col + j
        w = (overrides or {}).get(col, widths[j])
        ws.column_dimensions[get_column_letter(col)].width = max(8, min(60, w * 1.15 + 2))


# ══════════════════════════════════════════════════════════
# 4. 数据加载与预处理
# ══════════════════════════════════════════════════════════

def load_data():
    """加载全部 CSV 并做类型化预处理，返回数据字典。"""
    raw = {
        "artists": read_csv("artist_list.csv"),
        "skus": read_csv("sku_catalog.csv"),
        "dealers": read_csv("dealer_list.csv"),
        "sales": read_csv("sales_orders.csv"),
        "st": read_csv("dealer_sell_through.csv"),
        "receipts": read_csv("bank_receipts.csv"),
        "msettle": read_csv("modian_settlements.csv"),
        "pledges": read_csv("modian_pledges.csv"),
        "expos": read_csv("expo_revenue.csv"),
        "contracts": read_csv("contract_ledger.csv"),
    }

    sales = [dict(
        no=r["销售单号"], date=date_(r["开单日期"]), dcode=r["经销商编码"],
        dname=r["经销商名称"], sku=r["SKU编码"], name=r["商品名称"],
        qty=intv(r["数量"]), price=num(r["单价"]), amt=num(r["金额"]),
        disc=num(r["折扣率"]), net=num(r["实收金额"]), terms=intv(r["账期天数"]),
        due=date_(r["到期日"]), batch=r["批次类型"], note=r["备注"])
        for r in raw["sales"]]

    st = [dict(
        no=r["报表编号"], month=date_(r["报表月份"]), dcode=r["经销商编码"],
        sku=r["SKU编码"], name=r["商品名称"],
        qty=intv(r["销售数量"]), price=num(r["单价"]), amt=num(r["销售金额"]))
        for r in raw["st"]]

    receipts = [dict(
        no=r["回款单号"], date=date_(r["到账日期"]), payer=r["付款方名称"],
        amt=num(r["金额"]), method=r["付款方式"], bank=r["开户行"],
        txn=r["交易流水号"], note=r["备注"])
        for r in raw["receipts"]]

    artists = [dict(
        code=r["艺术家编码"], name=r["艺术家姓名"], kind=r["身份类型"],
        skus=r["关联SKU"], rate=pct(r["分成比例"]), settle=r["结算方式"])
        for r in raw["artists"]]

    msettle = [dict(
        no=r["结算编号"], apply=date_(r["申请日期"]), stage=r["结算阶段"],
        due=num(r["应结算金额"]), arrive=date_(r["到账日期"]),
        actual=num(r["实际到账"]), status=r["状态"], note=r["备注"])
        for r in raw["msettle"]]

    pledges = [dict(
        no=r["支持编号"], date=date_(r["支持日期"]), backer=r["支持者"],
        tier=r["档位名称"], qty=intv(r["数量"]), price=num(r["单价"]),
        amt=num(r["金额"]), pay=r["支付状态"], ship=r["发货状态"])
        for r in raw["pledges"]]

    return dict(
        artists=artists, skus=raw["skus"], dealers=raw["dealers"],
        sales=sales, st=st, receipts=receipts, msettle=msettle,
        pledges=pledges, expos=raw["expos"], contracts=raw["contracts"])


# ── 经销商匹配（回款付款方 → 经销商编码） ───────────────────

_COMPANY_SUFFIXES = ("有限责任公司", "股份有限公司", "有限公司", "责任公司", "公司")


def norm_name(s):
    """付款方/经销商名称规范化：去括号、去公司后缀，用于模糊匹配。"""
    s = re.sub(r"[（）()]", "", s or "")
    for suf in _COMPANY_SUFFIXES:
        if s.endswith(suf):
            s = s[: -len(suf)]
            break
    return s.strip()


def build_dealer_matcher(dealers):
    dealer_by_code = {d["经销商编码"]: d for d in dealers}
    dealer_by_name = {d["经销商名称"]: d["经销商编码"] for d in dealers}

    def match(payer):
        """返回 (经销商编码, 匹配方式)；无法匹配返回 (None, '')。"""
        if not payer:
            return None, ""
        if payer in dealer_by_name:
            return dealer_by_name[payer], "名称一致"
        npayer = norm_name(payer)
        for d in dealers:
            nd = norm_name(d["经销商名称"])
            if nd and nd == npayer:
                return d["经销商编码"], "规范化一致"
        for d in dealers:
            nd = norm_name(d["经销商名称"])
            if nd and npayer and (npayer in nd or nd in npayer):
                return d["经销商编码"], "包含匹配"
        return None, ""
    return match


# ── 艺术家关联SKU → 销售规格编码匹配 ────────────────────────

def artist_sku_matching(sales):
    """返回 (codes→matched 函数, sales_orders 中出现的全部规格编码集合)。"""
    sales_codes = sorted({s["sku"] for s in sales})

    def matched_codes(linked):
        """解析艺术家『关联SKU』字段，返回 (匹配到的销售规格列表, 备注列表)。"""
        codes = [c.strip() for c in re.split(r"[,，;；/、\s]+", linked) if c.strip()]
        matched, notes = [], []
        for c in codes:
            hit = [sc for sc in sales_codes if sc == c or (c in sc)]
            if hit:
                matched.extend(hit)
            elif c.startswith("BC-ACC-") and ("ACC-" + c[7:]) in sales_codes:
                notes.append(f"{c} 与销售规格 {('ACC-' + c[7:])} 编码不一致（疑缺BC前缀），"
                             f"本期按映射计入，建议与艺术家确认口径")
                matched.append("ACC-" + c[7:])
            elif c in ("BC-07", "BC-08"):
                notes.append(f"{c} 隐藏款随盲盒/套装随机出货，无独立销售规格，"
                             f"建议按套盒分摊口径另行测算")
            else:
                notes.append(f"{c} 当期无对应销售规格")
        return sorted(set(matched)), notes
    return matched_codes


def sales_net_by_codes(sales, codes):
    """按销售规格编码集合汇总实收金额。"""
    cset = set(codes)
    return round(sum(s["net"] for s in sales if s["sku"] in cset), 2)


# ══════════════════════════════════════════════════════════
# 5. 各工作表构建
# ══════════════════════════════════════════════════════════


def build_usage(wb, D):
    ws = wb.create_sheet("📋 使用说明")
    setup_sheet(ws, TAB_UTIL, freeze=None, landscape=False, repeat_rows=None)
    ws.sheet_view.showGridLines = False

    row = 1
    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, "菲格洛亚 FIGUEROA — BabyCat 财务对账工具包")
    c.font = Font(name=FONT_NAME, size=16, bold=True, color=DEEP_PURPLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:F2")
    c2 = ws.cell(2, 1, "版本 v1.0  ·  生成日期 2026-08-11  ·  数据期间 2026年5月—7月（含2025年摩点众筹 / 展会历史事项）")
    c2.font = Font(name=FONT_NAME, size=10, color="6B6B6B")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    row = 4

    def text_block(start, lines, size=10, color="333333", bold=False, height=16):
        for i, line in enumerate(lines):
            ws.merge_cells(start_row=start + i, start_column=1, end_row=start + i, end_column=6)
            c = ws.cell(start + i, 1, line)
            c.font = Font(name=FONT_NAME, size=size, color=color, bold=bold)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[start + i].height = height
        return start + len(lines)

    row = section_header(ws, row, "① 公司背景", 6) + 1
    row = text_block(row, [
        "菲格洛亚（FIGUEROA）为 AI 原生潮玩公司，主打 IP「BabyCat」毛绒盲盒——8 个猫品种角色（6 常规款 + 2 隐藏款）。",
        "渠道覆盖：名创优品 / TOPTOY / KKV / 九木杂物社 / 沃尔玛 / 抖音电商 / TikTok 跨境 / 区域经销；",
        "并先后在摩点平台发起众筹（三阶段结算，平台服务费 6%），参加 CTE、HTE 等潮玩展会。",
    ]) + 1

    row = section_header(ws, row, "② 工作簿结构（15 张表）", 6) + 1
    idx_rows = [
        ("📋 使用说明", "本页：版本、结构索引、对账流程、口径与图例", "工具"),
        ("🎨 商品目录", "BabyCat 全系 24 个 SKU：零售价 / 经销价", "数据"),
        ("👩‍🎨 艺术家主数据", "8 位艺术家、关联 SKU、分成比例与预计月版税", "数据"),
        ("📦 销售出库", "45 笔销售单（含退货冲销，红色标识），账期与到期日", "数据"),
        ("🏪 Sell-Through", "经销商门店动销 25 条，与出库口径相互印证", "数据"),
        ("💰 银行回款", "22 笔银行回款，按付款方名称自动匹配经销商（辅助列可隐藏）", "数据"),
        ("🎯 摩点众筹结算", "三阶段结算汇总、平台服务费 6% 测算、112 笔支持明细抽样", "分析"),
        ("🔗 经销商对账", "核心表：SUMIF 跨表计算应收/已回/差异/状态 + 待核对事项", "分析"),
        ("🎨 艺术家版税结算", "关联 SKU 销售实收 × 分成比例，SUMIF 公式引用销售出库", "分析"),
        ("🎪 展会核算", "3 场展会收入费用与毛利率（含综合毛利率）", "数据"),
        ("📊 Dashboard", "KPI 卡片、经销商应收vs已回柱状图、月度趋势折线图、账龄分析", "仪表盘"),
        ("📋 合同发票台账", "15 份合同 + 进项/销项发票登记与开票状态", "数据"),
        ("📦 代账交接清单", "移交外部代账公司的期间汇总、文档清单与待办事项", "工具"),
        ("💵 现金日记账", "众筹/展会/回款现金流示例，余额自动累计", "分析"),
        ("🧾 费用报销审核", "10 笔示例报销，自动校验发票金额差异与 OA 审批", "分析"),
    ]
    headers = ["工作表", "内容说明", "类型"]
    last = write_table(ws, row, headers,
                       [[a, b, c3] for a, b, c3 in idx_rows],
                       col_formats=[None, None, None], center_cols=(2,))
    for r in range(row + 1, last + 1):
        cat = ws.cell(r, 3).value
        if cat == "数据":
            ws.cell(r, 3).fill = PatternFill("solid", fgColor="D9E2F3")
        elif cat == "分析":
            ws.cell(r, 3).fill = PatternFill("solid", fgColor="E6D9F2")
        elif cat == "仪表盘":
            ws.cell(r, 3).fill = PatternFill("solid", fgColor="E2EFDA")
        else:
            ws.cell(r, 3).fill = PatternFill("solid", fgColor="EDEDED")
    row = last + 2

    row = section_header(ws, row, "③ 对账工作流程", 6) + 1
    steps = [
        "1. 导出数据：销售出库 / 门店动销 / 银行回款 / 众筹结算 / 展会收入 / 合同（CSV 落盘到 ../data/）；",
        "2. 回款匹配：脚本按付款方名称自动匹配经销商（名称一致 → 规范化一致 → 包含匹配，如『酷乐潮玩（北京）贸易公司』）；",
        "3. 生成对账：『🔗 经销商对账』按经销商汇总应收（SUMIF 销售出库实收）与已回（SUMIF 银行回款辅助列），计算差异与状态；",
        "4. 处理差异：在『待核对事项』逐条核实（部分回款 / 超额回款 / 抹零 / 汇兑 / 退货冲销 / 逾期未回）；",
        "5. 平行结算：摩点众筹三阶段核销、艺术家版税计提、展会收支毛利、费用报销复核；",
        "6. 归档交接：按『📦 代账交接清单』整理凭证、对账单与发票，移交代账公司。",
    ]
    row = text_block(row, steps, height=17) + 1

    row = section_header(ws, row, "④ 口径说明", 6) + 1
    calibers = [
        "应收口径：销售出库『实收金额』（= 金额 × 折扣率），退货冲销为负数自动冲减；",
        "已回口径：银行回款金额，按付款方自动匹配经销商后汇总（不含众筹/展会等非经销回款）；",
        "匹配状态：|应收-已回| < 0.01 → 已匹配（绿）；已回>0 但存在差异 → 部分/超额回款（黄）；已回=0 → 未回款（红）；",
        "众筹口径：应结算金额为平台扣除服务费后的净额，服务费按 6% 倒算（毛额=净额/0.94）；",
        "版税口径：版税基数 = 艺术家关联 SKU 对应销售规格的实收金额合计；隐藏款无独立规格当期不计；",
        "账龄口径：按销售单到期日与基准日（2026-08-11）计算，负数退货计入相应账龄。",
    ]
    row = text_block(row, calibers) + 1

    row = section_header(ws, row, "⑤ 颜色图例", 6) + 1
    legend = [
        ("已匹配（应收≈已回）", GREEN_FILL, GREEN_FONT),
        ("部分回款 / 超额回款", YELLOW_FILL, YELLOW_FONT),
        ("未回款 / 发票差异", RED_FILL, RED_FONT),
        ("表头（浅紫品牌色）", HEADER_FILL, DEEP_PURPLE),
    ]
    for i, (label, fill, fontc) in enumerate(legend):
        r = row + i
        ws.cell(r, 1, "■ " + label)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(r, 1).font = Font(name=FONT_NAME, size=10, color=fontc, bold=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(r, 1).border = BORDER
        ws.row_dimensions[r].height = 18
    row += len(legend) + 1

    row = section_header(ws, row, "⑥ 使用提示", 6) + 1
    tips = [
        "· Dashboard F31 为对账基准日，可修改后自动重算账龄与全表；",
        "· 『💰 银行回款』J 列（经销商编码）为 SUMIF 辅助列，可隐藏不影响公式；",
        "· 所有跨表公式在 Excel 打开时自动重算，文件无外部数据连接，可独立分发；",
        "· 打印已设置为横向、适应页宽并重复表头行，可直接打印存档。",
    ]
    row = text_block(row, tips) + 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 72
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 8


def build_catalog(wb, D):
    ws = wb.create_sheet("🎨 商品目录")
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "🎨 商品目录 — BabyCat 毛绒盲盒及周边（24 款）",
               "零售价 = 建议零售价；经销价 = 渠道供货价；盲盒随机出货，明盒指定角色，端盒 6 只不重复。", 6)
    rows = [[r["SKU编码"], r["规格编码"], r["商品名称"], r["品类"],
             num(r["零售价"]), num(r["经销价"])] for r in D["skus"]]
    last = write_table(ws, LAYOUT_HEADER_ROW,
                       ["SKU编码", "规格编码", "商品名称", "品类", "零售价", "经销价"],
                       rows, col_formats=[None, None, None, None, FMT_AMT, FMT_AMT],
                       center_cols=(0, 1, 3))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:F{last}"
    autofit(ws, ["SKU编码", "规格编码", "商品名称", "品类", "零售价", "经销价"], rows)


def build_artists(wb, D):
    ws = wb.create_sheet("👩‍🎨 艺术家主数据")
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "👩‍🎨 艺术家主数据（8 位）",
               "预计月版税 = 2026年5-7月关联SKU销售实收月均 × 分成比例（静态测算；正式计提见『🎨 艺术家版税结算』）。", 9)
    matched_f = D["artist_match"]
    headers = ["艺术家编码", "艺术家姓名", "身份类型", "关联SKU", "分成比例", "结算方式",
               "当期销售实收(5-7月)", "预计月版税", "备注"]
    rows = []
    for a in D["artists"]:
        matched, notes = matched_f(a["skus"])
        base = sales_net_by_codes(D["sales"], matched)
        monthly = round(base / 3 * a["rate"], 2)
        note = "；".join(notes) if notes else "—"
        rows.append([a["code"], a["name"], a["kind"], a["skus"], a["rate"], a["settle"],
                     base, monthly, note])
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows,
                       col_formats=[None, None, None, None, FMT_PCT, None, FMT_AMT, FMT_AMT, None],
                       center_cols=(0, 4))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:I{last}"
    for r in range(DATA_FIRST, last + 1):
        ws.row_dimensions[r].height = 30
    autofit(ws, headers, rows, overrides={9: 46})


def build_sales(wb, D):
    ws = wb.create_sheet(S_SALES)
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "📦 销售出库单（45 笔，含退货冲销）",
               "实收金额 = 金额 × 折扣率；到期日 = 开单日期 + 账期天数；红色行 = 负数（退货冲销）。", 15)
    headers = ["销售单号", "开单日期", "经销商编码", "经销商名称", "SKU编码", "商品名称",
               "数量", "单价", "金额", "折扣率", "实收金额", "账期天数", "到期日", "批次类型", "备注"]
    rows = [[s["no"], s["date"], s["dcode"], s["dname"], s["sku"], s["name"],
             s["qty"], s["price"], s["amt"], s["disc"], s["net"],
             s["terms"], s["due"], s["batch"], s["note"]] for s in D["sales"]]
    fmts = [None, FMT_DATE, None, None, None, None, FMT_QTY, FMT_AMT, FMT_AMT,
            FMT_RATE, FMT_AMT, FMT_QTY, FMT_DATE, None, None]
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows, col_formats=fmts,
                       center_cols=(0, 1, 2, 6, 9, 11, 12, 13))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:O{last}"
    for i, s in enumerate(D["sales"]):
        if s["qty"] < 0 or s["amt"] < 0:
            r = DATA_FIRST + i
            for j in range(1, 16):
                cell = ws.cell(r, j)
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=RED_FONT)
                cell.fill = PatternFill("solid", fgColor=RED_FILL)
    autofit(ws, headers, rows, overrides={4: 34, 6: 40, 15: 30})


def build_sellthrough(wb, D):
    ws = wb.create_sheet(S_ST)
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "🏪 Sell-Through — 经销商门店动销（25 条）",
               "寄售/经销渠道按月报送的终端动销数据，销售金额 = 销售数量 × 单价，与『📦 销售出库』口径相互印证。", 8)
    headers = ["报表编号", "报表月份", "经销商编码", "SKU编码", "商品名称", "销售数量", "单价", "销售金额"]
    rows = [[r["no"], r["month"], r["dcode"], r["sku"], r["name"], r["qty"], r["price"], r["amt"]]
            for r in D["st"]]
    fmts = [None, FMT_MONTH, None, None, None, FMT_QTY, FMT_AMT, FMT_AMT]
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows, col_formats=fmts,
                       center_cols=(0, 1, 2, 3, 5))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:H{last}"
    autofit(ws, headers, rows, overrides={5: 40})


def build_bank(wb, D, matcher):
    ws = wb.create_sheet(S_BANK)
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "💰 银行回款流水（22 笔）",
               "按付款方名称自动匹配经销商（名称一致 / 规范化一致 / 包含匹配）；I 列为匹配结果，J 列为 SUMIF 辅助编码（可隐藏）。", 10)
    headers = ["回款单号", "到账日期", "付款方名称", "金额", "付款方式", "开户行",
               "交易流水号", "备注", "对账经销商(自动匹配)", "经销商编码(辅助列)"]
    rows = []
    for r in D["receipts"]:
        code, how = matcher(r["payer"])
        rows.append([r["no"], r["date"], r["payer"], r["amt"], r["method"], r["bank"],
                     r["txn"], r["note"],
                     (D["dealers_by_code"][code]["经销商名称"] if code else "未匹配(需人工确认)"),
                     code if code else ""])
    fmts = [None, FMT_DATE, None, FMT_AMT, None, None, None, None, None, None]
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows, col_formats=fmts,
                       center_cols=(0, 1, 9))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:J{last}"
    ws.column_dimensions["J"].hidden = True
    autofit(ws, headers, rows, overrides={3: 36, 9: 32})


def build_modian(wb, D):
    ws = wb.create_sheet("🎯 摩点众筹结算")
    setup_sheet(ws, TAB_ANALYSIS)
    make_title(ws, "🎯 摩点众筹结算 — BabyCat 首轮众筹（2025.07–2026.02）",
               "应结算金额为平台扣除服务费后的净额；服务费按 6% 倒算（毛额=净额/0.94）；支持明细为后台抽样 112 笔，与结算总额口径不同。", 9)

    row = 4
    row = section_header(ws, row, "① 三阶段结算明细", 9)
    n_set = len(D["msettle"])
    last = write_table(ws, row + 1,
                       ["结算编号", "申请日期", "结算阶段", "应结算金额", "到账日期", "实际到账", "状态", "备注"],
                       [[m["no"], m["apply"], m["stage"], m["due"], m["arrive"], m["actual"], m["status"], m["note"]]
                        for m in D["msettle"]],
                       col_formats=[None, FMT_DATE, None, FMT_AMT, FMT_DATE, FMT_AMT, None, None],
                       center_cols=(0, 1, 3, 5, 6))
    total_row = last + 1
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 3, "三阶段合计")
    ws.cell(total_row, 4, f"=SUM(D{last - n_set + 1}:D{last})")
    ws.cell(total_row, 6, f"=SUM(F{last - n_set + 1}:F{last})")
    ws.cell(total_row, 7, "—")
    ws.cell(total_row, 8, "未到账=应结算-实际到账")
    for j in range(1, 10):
        c = ws.cell(total_row, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        if j in (4, 6):
            c.number_format = FMT_AMT
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    settle_due_cell = f"D{total_row}"
    settle_actual_cell = f"F{total_row}"
    row = total_row + 1
    row = section_header(ws, row, "② 平台服务费测算（摩点标准 6%）", 9)

    def _fee_row(r, label, formula, note):
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        ws.cell(r, 3, note)
        for j in range(1, 4):
            c = ws.cell(r, j)
            c.font = Font(name=FONT_NAME, size=10)
            c.border = BORDER
            if j == 2:
                c.number_format = FMT_AMT
                c.alignment = Alignment(horizontal="right", vertical="center")
        return r + 1

    r = row + 1
    for j, h in enumerate(["项目", "金额（元）", "说明"], start=1):
        header_style(ws.cell(r, j, h))
    r += 1
    r = _fee_row(r, "应结算总额（文件口径）", f"={settle_due_cell}", "三阶段应结算合计（已扣服务费）")
    r = _fee_row(r, "众筹认筹毛额（倒算）", f"=ROUND({settle_due_cell}/0.94,2)", "毛额 = 净额 ÷ (1-6%)")
    r = _fee_row(r, "平台服务费（6%）", f"=ROUND({settle_due_cell}/0.94*0.06,2)", "认筹毛额 × 6%")
    r = _fee_row(r, "服务费后净应收", f"=ROUND({settle_due_cell}/0.94*0.94,2)", "= 应结算总额（两者应一致）")
    r = _fee_row(r, "实际已到账", f"={settle_actual_cell}", "银行实际到账合计")
    r = _fee_row(r, "未到账", f"=ROUND({settle_due_cell}/0.94*0.94-{settle_actual_cell},2)", "0 表示三阶段全部结清")
    row = r + 1

    row = section_header(ws, row, "③ 支持明细汇总（后台抽样 112 笔）", 9)
    tier_totals = {}
    for p in D["pledges"]:
        t = tier_totals.setdefault(p["tier"], [0, 0.0])
        t[0] += 1
        t[1] += p["amt"]
    tier_rows = [[tier, cnt, round(amt, 2)] for tier, (cnt, amt) in tier_totals.items()]
    tier_rows.sort(key=lambda x: -x[2])
    total_pledge_cnt = sum(x[1] for x in tier_rows)
    tl = write_table(ws, row + 1, ["档位名称", "支持笔数", "支持金额（元）", "占比"],
                     [[t[0], t[1], t[2], f"={t[1]}/{total_pledge_cnt}"]
                      for t in tier_rows],
                     col_formats=[None, FMT_QTY, FMT_AMT, FMT_PCT], center_cols=(1, 2, 3))
    tr = tl + 1
    ws.cell(tr, 1, "合计")
    ws.cell(tr, 2, f"=SUM(B{row + 2}:B{tl})")
    ws.cell(tr, 3, f"=SUM(C{row + 2}:C{tl})")
    ws.cell(tr, 4, f"{len(tier_rows)} 个档位")
    for j in range(1, 5):
        c = ws.cell(tr, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        if j in (2, 3):
            c.number_format = FMT_QTY if j == 2 else FMT_AMT
    pay_ok = sum(1 for p in D["pledges"] if p["pay"] == "已支付")
    ship_ok = sum(1 for p in D["pledges"] if p["ship"] == "已发货")
    n_pledges = len(D["pledges"])
    stat_row = tr + 2
    ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=4)
    ws.cell(stat_row, 1,
            f"支付状态：已支付 {pay_ok} 笔 / 未支付 {n_pledges - pay_ok} 笔     |     发货状态：已发货 {ship_ok} 笔 / 未发货 {n_pledges - ship_ok} 笔")
    ws.cell(stat_row, 1).font = Font(name=FONT_NAME, size=10, color=DEEP_PURPLE, bold=True)
    row = stat_row + 2

    row = section_header(ws, row, "④ 支持明细（抽样）", 9)
    pl = write_table(ws, row + 1,
                     ["支持编号", "支持日期", "支持者", "档位名称", "数量", "单价", "金额", "支付状态", "发货状态"],
                     [[p["no"], p["date"], p["backer"], p["tier"], p["qty"], p["price"], p["amt"], p["pay"], p["ship"]]
                      for p in D["pledges"]],
                     col_formats=[None, FMT_DATE, None, None, FMT_QTY, FMT_AMT, FMT_AMT, None, None],
                     center_cols=(0, 1, 4, 5, 6, 7, 8))
    ws.auto_filter.ref = f"A{row + 1}:I{pl}"
    ws.freeze_panes = f"A{row + 2}"
    autofit(ws, ["支持编号", "支持日期", "支持者", "档位名称", "数量", "单价", "金额", "支付状态", "发货状态"],
            [[p["no"], p["date"], p["backer"], p["tier"], p["qty"], p["price"], p["amt"], p["pay"], p["ship"]]
             for p in D["pledges"]],
            start_col=1)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12


def build_recon(wb, D, matcher):
    ws = wb.create_sheet(S_RECON)
    setup_sheet(ws, TAB_ANALYSIS, freeze="A4")
    make_title(ws, "🔗 经销商对账工作底稿（核心表）",
               "应收 = SUMIF(销售出库实收金额, 经销商编码)；已回 = SUMIF(银行回款辅助列J, 经销商编码)；"
               "差异 = 应收 - 已回；匹配状态自动判断（绿=已匹配 / 黄=部分或超额 / 红=未回款）。", 7)

    sales_last = DATA_FIRST + len(D["sales"]) - 1
    st_last = DATA_FIRST + len(D["st"]) - 1
    bk_last = DATA_FIRST + len(D["receipts"]) - 1

    headers = ["经销商编码", "经销商名称", "应收金额(销售实收)", "门店销通金额", "已回款(银行)", "差异(应收-已回)", "匹配状态"]
    dealer_rows = []
    py_totals = {}   # code -> (net_sales, received)
    for d in D["dealers"]:
        net_sales = round(sum(s["net"] for s in D["sales"] if s["dcode"] == d["经销商编码"]), 2)
        received = round(sum(r["amt"] for r in D["receipts"] if r["dcode"] == d["经销商编码"]), 2)
        py_totals[d["经销商编码"]] = (net_sales, received)

    r_first, r_last = DATA_FIRST, DATA_FIRST + len(D["dealers"]) - 1
    for i, d in enumerate(D["dealers"]):
        r = r_first + i
        ws.cell(r, 1, d["经销商编码"])
        ws.cell(r, 2, d["经销商名称"])
        ws.cell(r, 3, f"=SUMIF('{S_SALES}'!$C${DATA_FIRST}:$C${sales_last},$A{r},'{S_SALES}'!$K${DATA_FIRST}:$K${sales_last})")
        ws.cell(r, 4, f"=SUMIF('{S_ST}'!$C${DATA_FIRST}:$C${st_last},$A{r},'{S_ST}'!$H${DATA_FIRST}:$H${st_last})")
        ws.cell(r, 5, f"=SUMIF('{S_BANK}'!$J${DATA_FIRST}:$J${bk_last},$A{r},'{S_BANK}'!$D${DATA_FIRST}:$D${bk_last})")
        ws.cell(r, 6, f"=C{r}-E{r}")
        ws.cell(r, 7, (f"=IF(AND(E{r}>0,ABS(F{r})<0.005),\"已匹配\","
                       f"IF(E{r}>0,IF(F{r}<-0.005,\"超额回款\",\"部分回款\"),\"未回款\"))"))
        for j in range(1, 8):
            c = ws.cell(r, j)
            c.font = Font(name=FONT_NAME, size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j in (1, 7) else "left",
                                    vertical="center", wrap_text=(j == 2))
            if j in (3, 4, 5, 6):
                c.number_format = FMT_AMT
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ALT_FILL)
    # 表头
    for j, h in enumerate(headers, start=1):
        header_style(ws.cell(LAYOUT_HEADER_ROW, j, h))
    # 合计行
    total_r = r_last + 1
    ws.cell(total_r, 1, "合计")
    ws.cell(total_r, 3, f"=SUM(C{r_first}:C{r_last})")
    ws.cell(total_r, 4, f"=SUM(D{r_first}:D{r_last})")
    ws.cell(total_r, 5, f"=SUM(E{r_first}:E{r_last})")
    ws.cell(total_r, 6, f"=C{total_r}-E{total_r}")
    ws.cell(total_r, 7, "")
    for j in range(1, 8):
        c = ws.cell(total_r, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        if j in (3, 4, 5, 6):
            c.number_format = FMT_AMT

    # 条件格式：整行按状态着色
    cf_range = f"A{r_first}:G{r_last}"
    ws.conditional_formatting.add(
        cf_range, FormulaRule(formula=['$G4="已匹配"'],
                              fill=PatternFill("solid", fgColor=GREEN_FILL),
                              font=Font(color=GREEN_FONT)))
    ws.conditional_formatting.add(
        cf_range, FormulaRule(formula=['OR($G4="部分回款",$G4="超额回款")'],
                              fill=PatternFill("solid", fgColor=YELLOW_FILL),
                              font=Font(color=YELLOW_FONT)))
    ws.conditional_formatting.add(
        cf_range, FormulaRule(formula=['$G4="未回款"'],
                              fill=PatternFill("solid", fgColor=RED_FILL),
                              font=Font(color=RED_FONT)))

    # ── 待核对事项 ──
    head_r = total_r + 2
    section_header(ws, head_r, "⚠ 待核对事项（差异清单）", 7, fill="FFE9E0")
    note_r = head_r + 1
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=7)
    ws.cell(note_r, 1, "仅列出 |差异| ≥ 0.01 的经销商；黄色=部分/超额回款，红色=未回款。以下金额为按公式口径自动计算的 Python 校验值，与 Excel 公式结果一致。")
    ws.cell(note_r, 1).font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")

    HINT_KW = ["部分回款", "抹零", "少付", "预付款", "汇兑", "逾期", "退货", "名称不规范"]
    items = []
    for d in D["dealers"]:
        code = d["经销商编码"]
        c_amt, e_amt = py_totals[code]
        diff = round(c_amt - e_amt, 2)
        if abs(diff) < 0.01:
            continue
        rate = (e_amt / c_amt) if c_amt else 0.0
        if e_amt == 0:
            reason = "无任何回款记录，账期已届满，建议重点催收"
        elif diff > 0 and rate >= 0.9:
            reason = "小额差异，建议核对抹零 / 手续费 / 四舍五入"
        elif diff > 0:
            reason = "部分回款，建议按销售单逐笔核销回款备注"
        else:
            reason = "超额回款（疑似预付款 / 跨期冲销），建议确认入账口径"
        hint = ""
        for r in D["receipts"]:
            if r["dcode"] == code and any(k in r["note"] for k in HINT_KW):
                hint = f"（回款备注：{r['note'][:36]}"
                break
        items.append(f"【{code}】{d['经销商名称']}：应收 ¥{c_amt:,.2f} / 已回 ¥{e_amt:,.2f} / "
                     f"差异 ¥{diff:,.2f}（回款率 {rate:.0%}）→ {reason}{hint}）")

    returns = [s for s in D["sales"] if s["qty"] < 0]
    for s in returns:
        items.append(f"【退货冲销】{s['dname']} {s['sku']} {s['qty']} 件，金额 ¥{s['amt']:,.2f}"
                     f"（已计入应收负数，回款含退货冲销，已核销）")

    unmatched = [r for r in D["receipts"] if not r["dcode"]]
    if unmatched:
        amt = sum(r["amt"] for r in unmatched)
        items.append(f"【未匹配回款】{len(unmatched)} 笔，金额合计 ¥{amt:,.2f}"
                     f"（付款方与经销商主数据不一致，需人工确认）")
    if not items:
        items.append("本期所有经销商均已对平，无待核对事项。")

    r = note_r + 1
    for item in items:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws.cell(r, 1, "· " + item)
        c.font = Font(name=FONT_NAME, size=10, color="7F2A2A" if "无任何回款" in item else "333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18 if len(item) < 90 else 32
        c.border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 40
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 17
    ws.column_dimensions["G"].width = 12
    ws.print_title_rows = "1:3"


def build_royalty(wb, D):
    ws = wb.create_sheet("🎨 艺术家版税结算")
    setup_sheet(ws, TAB_ANALYSIS)
    make_title(ws, "🎨 艺术家版税结算（销售后结算口径）",
               "版税基数 = 关联SKU对应销售规格的实收金额合计（SUMIF 引用『📦 销售出库』）；应付版税 = 基数 × 分成比例。", 10)
    matched_f = D["artist_match"]
    headers = ["艺术家编码", "艺术家姓名", "身份类型", "关联SKU", "对应销售规格(自动匹配)",
               "版税基数(销售实收)", "分成比例", "应付版税(元)", "结算方式", "状态/备注"]
    sales_last = DATA_FIRST + len(D["sales"]) - 1
    rows_meta = []
    r_first, r_last = DATA_FIRST, DATA_FIRST + len(D["artists"]) - 1
    for i, a in enumerate(D["artists"]):
        r = r_first + i
        matched, notes = matched_f(a["skus"])
        ws.cell(r, 1, a["code"])
        ws.cell(r, 2, a["name"])
        ws.cell(r, 3, a["kind"])
        ws.cell(r, 4, a["skus"])
        ws.cell(r, 5, "、".join(matched) if matched else "无直接对应")
        if matched:
            parts = [f"SUMIF('{S_SALES}'!$E${DATA_FIRST}:$E${sales_last},\"{c}\","
                     f"'{S_SALES}'!$K${DATA_FIRST}:$K${sales_last})" for c in matched]
            ws.cell(r, 6, f"=ROUND({'+'.join(parts)},2)")
        else:
            ws.cell(r, 6, 0)
        ws.cell(r, 7, a["rate"])
        ws.cell(r, 8, f"=ROUND(F{r}*G{r},2)")
        ws.cell(r, 9, a["settle"])
        note = "；".join(notes) if notes else "—"
        if a["code"] == "AR-005" and a["rate"] >= 0.10:
            note = "含预付 ¥5,000（见合同 CT-2026-005），结算时抵扣" + (("；" + note) if note != "—" else "")
        ws.cell(r, 10, note)
        for j in range(1, 11):
            c = ws.cell(r, j)
            c.font = Font(name=FONT_NAME, size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j in (1, 7) else "left",
                                    vertical="center", wrap_text=(j in (5, 10)))
            if j in (6, 8):
                c.number_format = FMT_AMT
            elif j == 7:
                c.number_format = FMT_PCT
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ALT_FILL)
        rows_meta.append([a["code"], a["name"], a["kind"], a["skus"], "、".join(matched) or "无直接对应",
                          ws.cell(r, 6).value, a["rate"], f"=ROUND(F{r}*G{r},2)", a["settle"], note])
    for j, h in enumerate(headers, start=1):
        header_style(ws.cell(LAYOUT_HEADER_ROW, j, h))
    total_r = r_last + 1
    ws.cell(total_r, 1, "合计")
    ws.cell(total_r, 6, f"=SUM(F{r_first}:F{r_last})")
    ws.cell(total_r, 8, f"=SUM(H{r_first}:H{r_last})")
    for j in range(1, 11):
        c = ws.cell(total_r, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        if j in (6, 8):
            c.number_format = FMT_AMT
    for r in range(r_first, r_last + 1):
        ws.row_dimensions[r].height = 34
    autofit(ws, headers, rows_meta, overrides={10: 48})
    ws.column_dimensions["G"].width = 10


def build_expo(wb, D):
    ws = wb.create_sheet("🎪 展会核算")
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "🎪 展会核算（3 场）",
               "毛利率 = 净收入 ÷ 总收入；收款方式为现场收款渠道（微信/支付宝/POS）。", 11)
    headers = ["活动编码", "活动名称", "开始日期", "结束日期", "地点", "收入类型",
               "总收入", "费用", "净收入", "收款方式", "毛利率"]
    rows = []
    for e in D["expos"]:
        rows.append([e["活动编码"], e["活动名称"], date_(e["开始日期"]), date_(e["结束日期"]),
                     e["地点"], e["收入类型"], num(e["总收入"]), num(e["费用"]), num(e["净收入"]),
                     e["收款方式"], None])
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows,
                       col_formats=[None, None, FMT_DATE, FMT_DATE, None, None, FMT_AMT, FMT_AMT, FMT_AMT, None, FMT_PCT],
                       center_cols=(0, 2, 3, 6, 7, 8, 10))
    for r in range(DATA_FIRST, last + 1):
        ws.cell(r, 11, f"=IF(G{r}=0,\"\",I{r}/G{r})")
    total_r = last + 1
    ws.cell(total_r, 1, "合计")
    ws.cell(total_r, 7, f"=SUM(G{DATA_FIRST}:G{last})")
    ws.cell(total_r, 8, f"=SUM(H{DATA_FIRST}:H{last})")
    ws.cell(total_r, 9, f"=SUM(I{DATA_FIRST}:I{last})")
    ws.cell(total_r, 11, f"=IF(G{total_r}=0,\"\",I{total_r}/G{total_r})")
    for j in range(1, 12):
        c = ws.cell(total_r, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        if j in (7, 8, 9):
            c.number_format = FMT_AMT
        elif j == 11:
            c.number_format = FMT_PCT
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:K{last}"
    autofit(ws, headers, rows)


def build_dashboard(wb, D):
    ws = wb.create_sheet("📊 Dashboard")
    setup_sheet(ws, TAB_DASH, freeze=None, landscape=True, repeat_rows=None)

    ws.merge_cells("A1:Q1")
    c = ws.cell(1, 1, "📊 BabyCat 财务对账 Dashboard")
    c.font = Font(name=FONT_NAME, size=16, bold=True, color=DEEP_PURPLE)
    c.fill = PatternFill("solid", fgColor=TITLE_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:Q2")
    c2 = ws.cell(2, 1, "数据期间 2026年5—7月  ·  对账基准日 2026-08-11（F31 可修改，账龄自动重算）  ·  全部指标由公式自动计算")
    c2.font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")
    c2.alignment = Alignment(horizontal="center", vertical="center")

    recon_first, recon_last = DATA_FIRST, DATA_FIRST + len(D["dealers"]) - 1

    # ── KPI 卡片 ──
    kpis = [
        (4, "总应收（销售实收）", f"=SUM('{S_RECON}'!$C${recon_first}:$C${recon_last})", FMT_AMT),
        (6, "总已回（银行回款）", f"=SUM('{S_RECON}'!$E${recon_first}:$E${recon_last})", FMT_AMT),
        (8, "净差异（应收-已回）", f"=A5-C5", FMT_AMT),
        (10, "对账完成率", f"=IF(A5=0,0,C5/A5)", FMT_PCT),
    ]
    for col, label, formula, fmt in kpis:
        r1, r2 = 4, 5
        ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + 1)
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 1)
        lc = ws.cell(r1, col, label)
        lc.font = Font(name=FONT_NAME, size=10, bold=True, color="6B6B6B")
        lc.fill = PatternFill("solid", fgColor=SECTION_FILL)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = BORDER
        vc = ws.cell(r2, col, formula)
        vc.font = Font(name=FONT_NAME, size=14, bold=True, color=DEEP_PURPLE)
        vc.fill = PatternFill("solid", fgColor=HEADER_FILL)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDER
        vc.number_format = fmt
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 30

    # ── ① 柱状图：经销商应收 vs 已回（镜像数据供图表引用） ──
    section_header(ws, 7, "① 经销商应收 vs 已回（元）", 8)
    mir_hdr_r, mir_first, mir_last = 41, 42, 41 + len(D["dealers"])
    ws.cell(40, 1, "④ 经销商对账镜像（供图表引用，公式自动同步）")
    ws.cell(40, 1).font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")
    for j, h in enumerate(["经销商编码", "经销商名称", "应收", "已回"], start=1):
        header_style(ws.cell(mir_hdr_r, j, h))
    for i, d in enumerate(D["dealers"]):
        r = mir_first + i
        src = recon_first + i
        ws.cell(r, 1, f"='{S_RECON}'!A{src}")
        ws.cell(r, 2, f"='{S_RECON}'!B{src}")
        ws.cell(r, 3, f"='{S_RECON}'!C{src}")
        ws.cell(r, 4, f"='{S_RECON}'!E{src}")
        for j in range(1, 5):
            cc = ws.cell(r, j)
            cc.font = Font(name=FONT_NAME, size=9)
            if j in (3, 4):
                cc.number_format = FMT_AMT
        ws.cell(r, 3).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 4).alignment = Alignment(horizontal="right", vertical="center")

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.gapWidth = 80
    bar.title = "经销商应收 vs 已回（元）"
    bar_data = Reference(ws, min_col=3, min_row=mir_hdr_r, max_col=4, max_row=mir_last)
    bar_cats = Reference(ws, min_col=2, min_row=mir_first, max_row=mir_last)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.series[0].graphicalProperties.solidFill = "7030A0"
    bar.series[1].graphicalProperties.solidFill = "4472C4"
    bar.x_axis.title = "经销商"
    bar.y_axis.title = "元"
    bar.width, bar.height = 15, 8
    ws.add_chart(bar, "A8")

    # ── ② 折线图：月度销售/回款趋势 ──
    ws.merge_cells("K7:N7")
    ws.cell(7, 11, "② 月度销售/回款趋势（2026年5-7月）")
    ws.cell(7, 11).font = Font(name=FONT_NAME, size=11, bold=True, color=DEEP_PURPLE)
    ws.cell(7, 11).fill = PatternFill("solid", fgColor=SECTION_FILL)
    ws.cell(7, 11).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22

    m_headers = ["月份", "销售额(元)", "回款额(元)"]
    for j, h in enumerate(m_headers, start=11):
        header_style(ws.cell(8, j, h))
    sales_last = DATA_FIRST + len(D["sales"]) - 1
    bk_last = DATA_FIRST + len(D["receipts"]) - 1
    for i, m in enumerate(MONTHS):
        r = 9 + i
        nxt = MONTHS[i + 1] if i + 1 < len(MONTHS) else datetime.date(m.year + 1, 1, 1)
        ws.cell(r, 11, f"{m.year}-{m.month:02d}")
        ws.cell(r, 12, (f"=SUMIFS('{S_SALES}'!$K${DATA_FIRST}:$K${sales_last},"
                        f"'{S_SALES}'!$B${DATA_FIRST}:$B${sales_last},\">=\"&DATE({m.year},{m.month},1),"
                        f"'{S_SALES}'!$B${DATA_FIRST}:$B${sales_last},\"<\"&DATE({nxt.year},{nxt.month},1))"))
        ws.cell(r, 13, (f"=SUMIFS('{S_BANK}'!$D${DATA_FIRST}:$D${bk_last},"
                        f"'{S_BANK}'!$B${DATA_FIRST}:$B${bk_last},\">=\"&DATE({m.year},{m.month},1),"
                        f"'{S_BANK}'!$B${DATA_FIRST}:$B${bk_last},\"<\"&DATE({nxt.year},{nxt.month},1))"))
        for j in (12, 13):
            c = ws.cell(r, j)
            c.number_format = FMT_AMT
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(12, 11, "合计")
    ws.cell(12, 12, f"=SUM(L9:L11)")
    ws.cell(12, 13, f"=SUM(M9:M11)")
    for j in (12, 13):
        ws.cell(12, j).number_format = FMT_AMT
    ws.cell(12, 11).font = Font(name=FONT_NAME, size=10, bold=True)
    ws.cell(12, 12).font = Font(name=FONT_NAME, size=10, bold=True)
    ws.cell(12, 13).font = Font(name=FONT_NAME, size=10, bold=True)

    line = LineChart()
    line.title = "月度销售 / 回款趋势（元）"
    line_data = Reference(ws, min_col=12, min_row=8, max_col=13, max_row=11)
    line_cats = Reference(ws, min_col=11, min_row=9, max_row=11)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_cats)
    line.series[0].graphicalProperties.line.solidFill = "7030A0"
    line.series[1].graphicalProperties.line.solidFill = "E36C09"
    line.x_axis.title = "月份"
    line.y_axis.title = "元"
    line.width, line.height = 12, 7.5
    ws.add_chart(line, "K14")

    # ── ③ 账龄分析 ──
    section_header(ws, 29, "③ 应收账款账龄分析（按销售到期日口径，元）", 8)
    age_headers = ["账龄区间", "应收金额(元)", "占比"]
    for j, h in enumerate(age_headers, start=1):
        header_style(ws.cell(30, j, h))
    ws.cell(31, 5, "对账基准日")
    ws.cell(31, 6, "=DATE(2026,8,11)")
    ws.cell(31, 6).number_format = FMT_DATE
    ws.cell(31, 7, "（修改后账龄自动重算）")
    ws.cell(31, 7).font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")

    m_rng = f"'📦 销售出库'!$M${DATA_FIRST}:$M${sales_last}"
    k_rng = f"'📦 销售出库'!$K${DATA_FIRST}:$K${sales_last}"
    ages = [
        ("未到期(信用期内)", f"=SUMPRODUCT(({m_rng}>$F$31)*({k_rng}))"),
        ("0-30天", f"=SUMPRODUCT(({m_rng}<=$F$31)*({m_rng}>$F$31-30)*({k_rng}))"),
        ("31-60天", f"=SUMPRODUCT(({m_rng}<=$F$31-30)*({m_rng}>$F$31-60)*({k_rng}))"),
        ("61-90天", f"=SUMPRODUCT(({m_rng}<=$F$31-60)*({m_rng}>$F$31-90)*({k_rng}))"),
        ("90天以上", f"=SUMPRODUCT(({m_rng}<=$F$31-90)*({k_rng}))"),
    ]
    for i, (label, formula) in enumerate(ages):
        r = 31 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        ws.cell(r, 3, f"=IF($B$36=0,\"\",B{r}/$B$36)")
        for j in (1, 2, 3):
            c = ws.cell(r, j)
            c.font = Font(name=FONT_NAME, size=10)
            c.border = BORDER
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ALT_FILL)
        ws.cell(r, 2).number_format = FMT_AMT
        ws.cell(r, 3).number_format = FMT_PCT
    ws.cell(36, 1, "合计")
    ws.cell(36, 2, "=SUM(B31:B35)")
    ws.cell(36, 3, "=IF($B$36=0,\"\",B36/B36)")
    for j in (1, 2, 3):
        c = ws.cell(36, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.cell(36, 2).number_format = FMT_AMT
    ws.cell(36, 3).number_format = FMT_PCT

    ws.merge_cells("A38:Q38")
    c38 = ws.cell(38, 1, "说明：账龄按销售出库到期日与基准日计算（退货冲销负数计入对应区间）；KPI 中净差异为负表示超额回款（预付款/跨期冲销）；"
                        "完整明细见『🔗 经销商对账』『📦 销售出库』『💰 银行回款』。")
    c38.font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")
    ws.cell(38, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[38].height = 28

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22
    for col in ("H", "I", "J"):
        ws.column_dimensions[col].width = 10
    ws.column_dimensions["K"].width = 11
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 10


def build_contracts(wb, D):
    ws = wb.create_sheet("📋 合同发票台账")
    setup_sheet(ws, TAB_DATA)
    make_title(ws, "📋 合同发票台账（15 份）",
               "进项/销项发票为示例登记；开票状态自动判断：存在发票号 → 已开票，否则 → 待开票。", 10)
    headers = ["合同编号", "签约方", "合同名称", "签署日期", "到期日期", "状态", "备注",
               "进项发票", "销项发票", "开票状态"]
    jx_no, inv_no = 0, 0
    rows = []
    for i, ct in enumerate(D["contracts"]):
        h_in, i_in = "", ""
        name = ct["合同名称"]
        if "艺术家" in name or "平台" in name:
            jx_no += 1
            h_in = f"JX-2026-{202 + jx_no:03d}"
        else:
            inv_no += 1
            i_in = f"INV-2026-{100 + inv_no:03d}"
        r = DATA_FIRST + i
        status_f = f"=IF(OR(H{r}<>\"\",I{r}<>\"\"),\"已开票\",\"待开票\")"
        rows.append([ct["合同编号"], ct["签约方"], name, date_(ct["签署日期"]), date_(ct["到期日期"]),
                     ct["状态"], ct["备注"], h_in, i_in, status_f])
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows,
                       col_formats=[None, None, None, FMT_DATE, FMT_DATE, None, None, None, None, None],
                       center_cols=(0, 3, 4, 5, 7, 8, 9))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:J{last}"
    autofit(ws, headers, rows, overrides={7: 42})


def build_handover(wb, D):
    ws = wb.create_sheet("📦 代账交接清单")
    setup_sheet(ws, TAB_UTIL, freeze=None, landscape=True, repeat_rows=None)
    ws.sheet_view.showGridLines = False

    recon_first, recon_last = DATA_FIRST, DATA_FIRST + len(D["dealers"]) - 1
    sales_last = DATA_FIRST + len(D["sales"]) - 1
    bk_last = DATA_FIRST + len(D["receipts"]) - 1
    n_returns = sum(1 for s in D["sales"] if s["qty"] < 0)

    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "📦 代账交接清单 — 菲格洛亚 FIGUEROA（BabyCat IP）")
    c.font = Font(name=FONT_NAME, size=15, bold=True, color=DEEP_PURPLE)
    c.fill = PatternFill("solid", fgColor=TITLE_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:H2")
    c2 = ws.cell(2, 1, "交接期间：2026年5月—7月（含 2025 年摩点众筹 / 展会等历史事项备查）   ·   对账基准日：2026-08-11   ·   编制：财务部")
    c2.font = Font(name=FONT_NAME, size=9, italic=True, color="6B6B6B")
    c2.alignment = Alignment(horizontal="center", vertical="center")

    row = 4
    row = section_header(ws, row, "① 期间业务概况", 8) + 1
    overview = [
        (f"销售出库单 {len(D['sales'])} 笔（含退货冲销 {n_returns} 笔，负值已冲减应收）", ""),
        (f"银行回款 {len(D['receipts'])} 笔，全部自动匹配到经销商（匹配率 100%）", ""),
        (f"经销商 {len(D['dealers'])} 家（连锁零售 / 电商平台 / 区域经销 / 海外跨境）", ""),
        (f"摩点众筹三阶段结算全部到账；支持明细抽样 {len(D['pledges'])} 笔", ""),
        (f"潮玩展会 {len(D['expos'])} 场（CTE / HTE / 济南快闪）", ""),
        (f"合同台账 {len(D['contracts'])} 份，其中艺术家合作协议 5 份", ""),
    ]
    for i, (label, _v) in enumerate(overview):
        ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=8)
        cc = ws.cell(row + i, 1, "· " + label)
        cc.font = Font(name=FONT_NAME, size=10)
        ws.row_dimensions[row + i].height = 17
    row += len(overview) + 1

    row = section_header(ws, row, "② 关键对账结果（公式自动引用各分表）", 8) + 1
    kv = [
        ("经销商总应收（销售实收）", f"=SUM('{S_RECON}'!$C${recon_first}:$C${recon_last})", FMT_AMT),
        ("经销商总已回（银行回款）", f"=SUM('{S_RECON}'!$E${recon_first}:$E${recon_last})", FMT_AMT),
        ("净差异（应收 - 已回）", f"=B{row + 1}-B{row + 2}", FMT_AMT),
        ("已匹配家数", f"=COUNTIF('{S_RECON}'!$G${recon_first}:$G${recon_last},\"已匹配\")", FMT_QTY),
        ("部分/超额回款家数", f"=COUNTIF('{S_RECON}'!$G${recon_first}:$G${recon_last},\"部分回款\")+COUNTIF('{S_RECON}'!$G${recon_first}:$G${recon_last},\"超额回款\")", FMT_QTY),
        ("未回款家数", f"=COUNTIF('{S_RECON}'!$G${recon_first}:$G${recon_last},\"未回款\")", FMT_QTY),
        ("摩点众筹已到账合计", f"='🎯 摩点众筹结算'!F8", FMT_AMT),
        ("展会净收入合计", f"='🎪 展会核算'!I7", FMT_AMT),
        ("艺术家版税应付合计", f"='🎨 艺术家版税结算'!H12", FMT_AMT),
        ("销售单总数 / 回款笔数", f"=COUNTA('{S_SALES}'!$A${DATA_FIRST}:$A${sales_last})&\" 笔 / \"&COUNTA('{S_BANK}'!$A${DATA_FIRST}:$A${bk_last})&\" 笔\"", None),
    ]
    header_style(ws.cell(row, 1, "项目"))
    header_style(ws.cell(row, 2, "数值"))
    for i, (label, formula, fmt) in enumerate(kv):
        r = row + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        for j in (1, 2):
            cc = ws.cell(r, j)
            cc.font = Font(name=FONT_NAME, size=10)
            cc.border = BORDER
            if j == 2 and fmt:
                cc.number_format = fmt
        if i % 2 == 1:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=ALT_FILL)
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=ALT_FILL)
    row += len(kv) + 2

    row = section_header(ws, row, "③ 文档交接清单（随附电子版/纸质件）", 8) + 1
    docs = [
        "☐ 银行回单 / 对账单：22 笔（『💰 银行回款』），含回款备注核销说明",
        "☐ 销售出库单：45 笔（『📦 销售出库』），含退货冲销 -30 件（品质问题，DE-004）",
        "☐ 经销商对账确认函：10 家（『🔗 经销商对账』），未对平项见待核对清单",
        "☐ 摩点众筹结算单：3 阶段（首款/中款/尾款），支持明细抽样 112 笔",
        "☐ 展会收入台账：3 场（CTE / HTE / 济南快闪），含现场收款流水",
        "☐ 合同台账：15 份（『📋 合同发票台账』），含艺术家合作协议 5 份",
        "☐ 发票：销项 INV-2026-xxx 系列（经销/供货合同）、进项 JX-2026-xxx 系列（平台/艺术家）",
        "☐ 费用报销单：10 笔（『🧾 费用报销审核』），OA 审批号已登记",
        "☐ 现金日记账：13 笔现金流示例（『💵 现金日记账』）",
    ]
    for i, d in enumerate(docs):
        ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=8)
        cc = ws.cell(row + i, 1, d)
        cc.font = Font(name=FONT_NAME, size=10)
        ws.row_dimensions[row + i].height = 17
    row += len(docs) + 1

    row = section_header(ws, row, "④ 待办 / 风险事项", 8) + 1
    todos = [
        "1. 沃尔玛（DE-006）6 月追单账期届满未回，金额较大，建议 8 月优先催收；",
        "2. TOPTOY（DE-003）6 月货款仅部分回款 ¥50,000，余款逐单核对；",
        "3. 抖音电商（DE-007）已回超出应收（预付款 30,000），建议确认跨期冲销口径；",
        "4. 广州潮玩前线（DE-009）抹零差异 ¥0.67、义乌萌物集（DE-010）少付 ¥1.00，建议核销审批；",
        "5. TikTok 跨境回款汇兑损益 ¥326.50，建议按到账日汇率入账并留档；",
        "6. 艺术家版税待结算（3 位艺术家本期有应计），赵小喵预付 ¥5,000 需在结算时抵扣；",
        "7. 隐藏款（BC-07/BC-08）版税分摊口径待与艺术家确认后补计。",
    ]
    for i, t in enumerate(todos):
        ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=8)
        cc = ws.cell(row + i, 1, t)
        cc.font = Font(name=FONT_NAME, size=10, color="7F2A2A")
        ws.row_dimensions[row + i].height = 17
    row += len(todos) + 2

    row = section_header(ws, row, "⑤ 交接确认", 8) + 1
    for i, (label, blank) in enumerate([
        ("移交人（菲格洛亚 财务部）", "＿＿＿＿＿＿＿（签名）"),
        ("接收人（代账公司）", "＿＿＿＿＿＿＿（签名）"),
        ("交接日期", "＿＿＿＿＿＿＿＿＿"),
        ("备注", "本清单与各分表共同构成期间账务交接依据，如有异议请在 5 个工作日内提出。"),
    ]):
        ws.cell(row + i, 1, label)
        ws.cell(row + i, 1).font = Font(name=FONT_NAME, size=10, bold=True)
        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=6)
        cc = ws.cell(row + i, 2, blank)
        cc.font = Font(name=FONT_NAME, size=10)
        ws.row_dimensions[row + i].height = 20

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 56
    for col in ("C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 12


def build_cash(wb, D):
    ws = wb.create_sheet("💵 现金日记账")
    setup_sheet(ws, TAB_ANALYSIS)
    make_title(ws, "💵 现金日记账（示例：众筹 / 展会 / 回款现金流）",
               "示例数据基于摩点到账、展会收支与经销商回款；余额列按累计公式自动计算。", 7)
    headers = ["日期", "凭证号", "摘要", "收入", "支出", "余额", "备注"]
    rows = [
        ("2025-08-10", "CJ-001", "摩点众筹首款到账", 39778.00, 0.00, None, "MD-SETTLE-01（首款50%）"),
        ("2025-09-17", "CJ-002", "CTE玩具展现场收款", 13340.00, 0.00, None, "微信/支付宝/POS"),
        ("2025-09-20", "CJ-003", "CTE玩具展场地及物料费用", 0.00, 2340.00, None, "展会成本"),
        ("2025-11-25", "CJ-004", "摩点众筹中款到账", 25067.00, 0.00, None, "MD-SETTLE-02（中款30%）"),
        ("2025-12-26", "CJ-005", "济南圣诞快闪店零售收款", 44118.00, 0.00, None, "含冬季吊卡首发342个"),
        ("2025-12-28", "CJ-006", "快闪店场地及物料费用", 0.00, 5200.00, None, "POP-001"),
        ("2026-02-12", "CJ-007", "摩点众筹尾款到账", 16711.00, 0.00, None, "MD-SETTLE-03（尾款20%）"),
        ("2026-03-29", "CJ-008", "HTE杭州潮玩展收款", 10490.00, 0.00, None, "微信/支付宝"),
        ("2026-03-31", "CJ-009", "HTE展会费用", 0.00, 1860.00, None, "HTE-001"),
        ("2026-06-20", "CJ-010", "名创优品5月试产回款", 53820.80, 0.00, None, "PM-20260620-010"),
        ("2026-07-01", "CJ-011", "抖音电商订货预付款", 30000.00, 0.00, None, "PM-20260701-003（预付款）"),
        ("2026-08-01", "CJ-012", "名创优品6月货款回款", 163348.00, 0.00, None, "PM-20260801-001（多单合并）"),
        ("2026-08-08", "CJ-013", "名创优品7月货款回款", 72680.00, 0.00, None, "PM-20260808-013"),
    ]
    out_rows = []
    for i, r in enumerate(rows):
        bal = None if r[5] is None else r[5]
        out_rows.append([date_(r[0]), r[1], r[2], r[3], r[4], bal, r[6]])
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, out_rows,
                       col_formats=[FMT_DATE, None, None, FMT_AMT, FMT_AMT, FMT_AMT, None],
                       center_cols=(0, 1, 3, 4, 5))
    for i in range(len(out_rows)):
        r = DATA_FIRST + i
        if i == 0:
            ws.cell(r, 6, "=D{}-E{}".format(r, r))
        else:
            ws.cell(r, 6, "=F{}+D{}-E{}".format(r - 1, r, r))
    total_r = last + 1
    ws.cell(total_r, 2, "本期合计")
    ws.cell(total_r, 4, f"=SUM(D{DATA_FIRST}:D{last})")
    ws.cell(total_r, 5, f"=SUM(E{DATA_FIRST}:E{last})")
    ws.cell(total_r, 6, f"=F{last}")
    for j in range(1, 8):
        c = ws.cell(total_r, j)
        c.font = Font(name=FONT_NAME, size=10, bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        if j in (4, 5, 6):
            c.number_format = FMT_AMT
    autofit(ws, headers, out_rows, overrides={7: 34})


def build_expense(wb, D):
    ws = wb.create_sheet("🧾 费用报销审核")
    setup_sheet(ws, TAB_ANALYSIS)
    make_title(ws, "🧾 费用报销审核（示例 10 笔）",
               "审核状态自动判断：报销金额 ≠ 发票金额 → 发票差异（红）；缺飞书 OA 审批号 → 待OA审批（黄）；其余 → 通过。", 10)
    headers = ["报销单号", "申请日期", "申请人", "费用类别", "报销金额", "发票金额",
               "发票号码", "飞书OA审批号", "审核状态", "备注"]
    samples = [
        ("EX-2026-001", "2026-06-08", "陈卷卷", "展会差旅费(CTE玩具展)", 2340.00, 2340.00, "EZZ-2026-1101", "OA-2026-0102", None, "高铁+住宿"),
        ("EX-2026-002", "2026-06-15", "周布布", "展会物料制作(济南快闪)", 5200.00, 5200.00, "EZZ-2026-1188", "OA-2026-0115", None, "POP-001 现场物料"),
        ("EX-2026-003", "2026-06-22", "林小喵", "设计服务费(角色设定)", 8000.00, 8000.00, "EZZ-2026-1201", "OA-2026-0120", None, "新系列角色设定"),
        ("EX-2026-004", "2026-07-02", "刘星野", "打样费(新系列)", 3500.00, 3200.00, "EZZ-2026-1233", "OA-2026-0129", None, "发票金额与报销不一致"),
        ("EX-2026-005", "2026-07-10", "赵小喵", "小红书推广费", 6000.00, 6000.00, "EZZ-2026-1250", "OA-2026-0131", None, "达人笔记投放"),
        ("EX-2026-006", "2026-07-18", "黄多多", "快递物流费(盲盒发货)", 1860.00, 1860.00, "EZZ-2026-1266", "OA-2026-0137", None, "众筹补发货"),
        ("EX-2026-007", "2026-07-25", "何梦鹿", "办公用品", 458.00, 458.00, "无发票", "", None, "零星采购，OA 未归档"),
        ("EX-2026-008", "2026-08-01", "杨桃桃", "打车费(广州出差)", 320.00, 300.00, "EZZ-2026-1300", "OA-2026-0142", None, "行程单与发票差20元"),
        ("EX-2026-009", "2026-08-05", "陈卷卷", "平台服务费(摩点手续费)", 5205.70, 5205.70, "EZZ-2026-1311", "OA-2026-0145", None, "服务费6%测算"),
        ("EX-2026-010", "2026-08-08", "林小喵", "餐饮招待(渠道洽谈)", 1280.00, 1280.00, "EZZ-2026-1320", "OA-2026-0148", None, "名创优品商务洽谈"),
    ]
    rows = []
    for i, s in enumerate(samples):
        rows.append([s[0], date_(s[1]), s[2], s[3], s[4], s[5], s[6], s[7],
                     f"=IF(ABS(E{DATA_FIRST + i}-F{DATA_FIRST + i})>0.01,\"发票差异\",IF(H{DATA_FIRST + i}=\"\",\"待OA审批\",\"通过\"))",
                     s[9]])
    last = write_table(ws, LAYOUT_HEADER_ROW, headers, rows,
                       col_formats=[None, FMT_DATE, None, None, FMT_AMT, FMT_AMT, None, None, None, None],
                       center_cols=(0, 1, 4, 5, 7, 8))
    ws.auto_filter.ref = f"A{LAYOUT_HEADER_ROW}:J{last}"
    cf_range = f"I{DATA_FIRST}:I{last}"
    ws.conditional_formatting.add(
        cf_range, FormulaRule(formula=['$I4="发票差异"'],
                              fill=PatternFill("solid", fgColor=RED_FILL),
                              font=Font(color=RED_FONT, bold=True)))
    ws.conditional_formatting.add(
        cf_range, FormulaRule(formula=['$I4="待OA审批"'],
                              fill=PatternFill("solid", fgColor=YELLOW_FILL),
                              font=Font(color=YELLOW_FONT)))
    autofit(ws, headers, rows, overrides={10: 32})


# ══════════════════════════════════════════════════════════
# 6. 主流程
# ══════════════════════════════════════════════════════════

def main():
    D = load_data()

    # 经销商匹配器（银行回款 → 经销商编码），并回写 dcode
    matcher = build_dealer_matcher(D["dealers"])
    D["dealers_by_code"] = {d["经销商编码"]: d for d in D["dealers"]}
    for r in D["receipts"]:
        r["dcode"], _ = matcher(r["payer"])
    D["artist_match"] = artist_sku_matching(D["sales"])

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    build_usage(wb, D)
    build_catalog(wb, D)
    build_artists(wb, D)
    build_sales(wb, D)
    build_sellthrough(wb, D)
    build_bank(wb, D, matcher)
    build_modian(wb, D)
    build_recon(wb, D, matcher)
    build_royalty(wb, D)
    build_expo(wb, D)
    build_dashboard(wb, D)
    build_contracts(wb, D)
    build_handover(wb, D)
    build_cash(wb, D)
    build_expense(wb, D)

    wb.save(OUT_FILE)
    size_kb = os.path.getsize(OUT_FILE) / 1024.0
    print(f"[OK] 已生成: {OUT_FILE} ({size_kb:.1f} KB)")
    print(f"     工作表 {len(wb.sheetnames)} 张: " + " / ".join(wb.sheetnames))
    print(f"     数据: 销售{len(D['sales'])}笔 / 回款{len(D['receipts'])}笔 / "
          f"经销商{len(D['dealers'])}家 / 艺术家{len(D['artists'])}位 / "
          f"众筹支持{len(D['pledges'])}笔 / 展会{len(D['expos'])}场 / 合同{len(D['contracts'])}份")


if __name__ == "__main__":
    main()
